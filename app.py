from flask import Flask, render_template, request, redirect, url_for, session, send_file
from database import init_db
from datetime import datetime, timedelta
import sqlite3, io

app = Flask(__name__)
app.secret_key = "cet22paniol2024"
DB_PATH = "paniol_cet22.db"
USUARIO = "admin"
CLAVE   = "cet22"

CURSOS = [
    "1°1 CB","1°2 CB","1°3 CB",
    "2°1 CB","2°2 CB","2°3 CB","2°4 CB",
    "1°1 CS","1°2 CS","1°3 CS","1°4 CS",
    "2°1 CS","2°2 CS","2°3 CS","2°4 CS",
    "3°1 CS","3°2 CS","3°3 CS",
    "4°1 CS","4°2 CS"
]

TALLERES = [
    "Ajuste I","Electricidad I","Carpinteria","Hojalateria",
    "Ajuste II","Electricidad II","Herreria y Soldadura","Informatica",
    "Electricidad III","Soldadura I","Oficina I","Maq. y Herr. I",
    "Soldaduras Especiales","Electricidad Industrial","Oficina II",
    "Maq. y Herr. II","Oficina III","Programacion CNC",
    "Automatizacion","Metalurgica Aplicada","Maq. y Herr. CNC"
]

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def login_req():
    return not session.get("logged")

# ── LOGIN ─────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    err = None
    if request.method == "POST":
        if request.form["user"] == USUARIO and request.form["clave"] == CLAVE:
            session["logged"] = True
            return redirect(url_for("index"))
        err = "Usuario o clave incorrectos"
    return render_template("login.html", error=err)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── INICIO ────────────────────────────────────────
@app.route("/")
def index():
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    prestamos = conn.execute("""
        SELECT p.id, h.nombre as herramienta, u.nombre as usuario,
               u.curso, u.turno, p.sector, p.destino, p.fecha_prestamo,
               p.observaciones, p.fecha_devolucion
        FROM prestamos p
        JOIN herramientas h ON p.herramienta_id = h.id
        JOIN usuarios u ON p.usuario_id = u.id
        WHERE p.fecha_devolucion IS NULL
        ORDER BY p.fecha_prestamo DESC
    """).fetchall()
    conn.close()
    return render_template("index.html", prestamos=prestamos)

# ── HERRAMIENTAS ──────────────────────────────────
@app.route("/herramientas")
def herramientas():
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    items = conn.execute("SELECT * FROM herramientas ORDER BY nombre").fetchall()
    conn.close()
    return render_template("herramientas.html", herramientas=items)

@app.route("/herramientas/nueva", methods=["GET","POST"])
def nueva_herramienta():
    if login_req(): return redirect(url_for("login"))
    if request.method == "POST":
        conn = get_conn()
        conn.execute("INSERT INTO herramientas (nombre,descripcion,cantidad,estado) VALUES (?,?,?,?)",
            (request.form["nombre"], request.form["descripcion"],
             request.form["cantidad"], request.form["estado"]))
        conn.commit(); conn.close()
        return redirect(url_for("herramientas"))
    return render_template("herramienta_form.html", herramienta=None)

@app.route("/herramientas/editar/<int:id>", methods=["GET","POST"])
def editar_herramienta(id):
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    if request.method == "POST":
        conn.execute("UPDATE herramientas SET nombre=?,descripcion=?,cantidad=?,estado=? WHERE id=?",
            (request.form["nombre"], request.form["descripcion"],
             request.form["cantidad"], request.form["estado"], id))
        conn.commit(); conn.close()
        return redirect(url_for("herramientas"))
    h = conn.execute("SELECT * FROM herramientas WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("herramienta_form.html", herramienta=h)

@app.route("/herramientas/eliminar/<int:id>")
def eliminar_herramienta(id):
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    conn.execute("DELETE FROM herramientas WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("herramientas"))

@app.route("/herramientas/importar", methods=["GET","POST"])
def importar_herramientas():
    if login_req(): return redirect(url_for("login"))
    msg = None
    if request.method == "POST":
        f = request.files.get("archivo")
        if f:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                conn = get_conn()
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre, descripcion, cantidad, estado = (list(row) + [None,None,None,None])[:4]
                    if nombre:
                        conn.execute("INSERT INTO herramientas (nombre,descripcion,cantidad,estado) VALUES (?,?,?,?)",
                            (str(nombre), str(descripcion or ""), int(cantidad or 1), str(estado or "Bueno")))
                        count += 1
                conn.commit(); conn.close()
                msg = f"Se importaron {count} herramientas correctamente."
            except Exception as e:
                msg = f"Error al importar: {e}"
    return render_template("importar_herramientas.html", msg=msg)

# ── USUARIOS ──────────────────────────────────────
@app.route("/usuarios")
def usuarios():
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    items = conn.execute("SELECT * FROM usuarios ORDER BY nombre").fetchall()
    conn.close()
    return render_template("usuarios.html", usuarios=items)

@app.route("/usuarios/nuevo", methods=["GET","POST"])
def nuevo_usuario():
    if login_req(): return redirect(url_for("login"))
    if request.method == "POST":
        conn = get_conn()
        conn.execute("INSERT INTO usuarios (nombre,curso,division,turno,tipo) VALUES (?,?,?,?,?)",
            (request.form["nombre"], request.form["curso"],
             request.form["division"], request.form["turno"], request.form["tipo"]))
        conn.commit(); conn.close()
        return redirect(url_for("usuarios"))
    return render_template("usuario_form.html", usuario=None)

@app.route("/usuarios/editar/<int:id>", methods=["GET","POST"])
def editar_usuario(id):
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    if request.method == "POST":
        conn.execute("UPDATE usuarios SET nombre=?,curso=?,division=?,turno=?,tipo=? WHERE id=?",
            (request.form["nombre"], request.form["curso"],
             request.form["division"], request.form["turno"], request.form["tipo"], id))
        conn.commit(); conn.close()
        return redirect(url_for("usuarios"))
    u = conn.execute("SELECT * FROM usuarios WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("usuario_form.html", usuario=u)

@app.route("/usuarios/eliminar/<int:id>")
def eliminar_usuario(id):
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    conn.execute("DELETE FROM usuarios WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("usuarios"))

@app.route("/usuarios/importar", methods=["GET","POST"])
def importar_usuarios():
    if login_req(): return redirect(url_for("login"))
    msg = None
    if request.method == "POST":
        f = request.files.get("archivo")
        if f:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                conn = get_conn()
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre, curso, division, turno, tipo = (list(row) + [None,None,None,None,None])[:5]
                    if nombre:
                        conn.execute("INSERT INTO usuarios (nombre,curso,division,turno,tipo) VALUES (?,?,?,?,?)",
                            (str(nombre), str(curso or ""), str(division or ""),
                             str(turno or "Manana"), str(tipo or "Alumno")))
                        count += 1
                conn.commit(); conn.close()
                msg = f"Se importaron {count} usuarios correctamente."
            except Exception as e:
                msg = f"Error al importar: {e}"
    return render_template("importar_usuarios.html", msg=msg)

# ── PRESTAMOS ─────────────────────────────────────
@app.route("/prestamos/nuevo", methods=["GET","POST"])
def nuevo_prestamo():
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    if request.method == "POST":
        sector  = request.form["sector"]
        destino = request.form.get("destino_curso") if sector == "Aula" else request.form.get("destino_taller")
        conn.execute("""INSERT INTO prestamos
            (herramienta_id,usuario_id,sector,destino,observaciones,fecha_prestamo)
            VALUES (?,?,?,?,?,?)""",
            (request.form["herramienta_id"], request.form["usuario_id"],
             sector, destino, request.form["observaciones"],
             datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit(); conn.close()
        return redirect(url_for("index"))
    herramientas = conn.execute("SELECT * FROM herramientas WHERE cantidad>0 ORDER BY nombre").fetchall()
    usuarios = conn.execute("SELECT * FROM usuarios ORDER BY nombre").fetchall()
    conn.close()
    return render_template("prestamo_form.html",
        herramientas=herramientas, usuarios=usuarios,
        cursos=CURSOS, talleres=TALLERES)

@app.route("/prestamos/devolver/<int:id>")
def devolver(id):
    if login_req(): return redirect(url_for("login"))
    conn = get_conn()
    conn.execute("UPDATE prestamos SET fecha_devolucion=? WHERE id=?",
                 (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), id))
    conn.commit(); conn.close()
    return redirect(url_for("index"))

# ── BUSCAR ────────────────────────────────────────
@app.route("/buscar")
def buscar():
    if login_req(): return redirect(url_for("login"))
    q      = request.args.get("q", "")
    turno  = request.args.get("turno", "todos")
    sector = request.args.get("sector", "todos")
    conn = get_conn()
    sql = """SELECT p.id, h.nombre as herramienta, u.nombre as usuario,
             u.curso, u.division, u.turno, u.tipo, p.sector, p.destino,
             p.fecha_prestamo, p.fecha_devolucion, p.observaciones
             FROM prestamos p
             JOIN herramientas h ON p.herramienta_id = h.id
             JOIN usuarios u ON p.usuario_id = u.id
             WHERE (h.nombre LIKE ? OR u.nombre LIKE ?)"""
    params = [f"%{q}%", f"%{q}%"]
    if turno  != "todos": sql += " AND u.turno=?";  params.append(turno)
    if sector != "todos": sql += " AND p.sector=?"; params.append(sector)
    sql += " ORDER BY p.fecha_prestamo DESC"
    resultados = conn.execute(sql, params).fetchall()
    conn.close()
    return render_template("buscar.html", resultados=resultados, q=q, turno=turno, sector=sector)

# ── INFORMES ──────────────────────────────────────
@app.route("/informes")
def informes():
    if login_req(): return redirect(url_for("login"))
    return render_template("informes.html")

def get_prestamos_informe(tipo, turno):
    hoy = datetime.now()
    if   tipo == "turno":         fecha_desde = hoy.replace(hour=0, minute=0, second=0)
    elif tipo == "semanal":       fecha_desde = hoy - timedelta(days=7)
    elif tipo == "mensual":       fecha_desde = hoy - timedelta(days=30)
    elif tipo == "cuatrimestral": fecha_desde = hoy - timedelta(days=120)
    else:                         fecha_desde = hoy - timedelta(days=365)
    conn = get_conn()
    sql = """SELECT p.id, h.nombre as herramienta, h.estado as estado_herramienta,
             u.nombre as usuario, u.curso, u.division, u.turno, u.tipo,
             p.sector, p.destino, p.fecha_prestamo, p.fecha_devolucion, p.observaciones
             FROM prestamos p
             JOIN herramientas h ON p.herramienta_id = h.id
             JOIN usuarios u ON p.usuario_id = u.id
             WHERE p.fecha_prestamo >= ?"""
    params = [fecha_desde.strftime("%Y-%m-%d %H:%M:%S")]
    if turno != "todos": sql += " AND u.turno=?"; params.append(turno)
    sql += " ORDER BY p.fecha_prestamo DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows, fecha_desde, hoy

@app.route("/informes/generar")
def generar_informe():
    if login_req(): return redirect(url_for("login"))
    tipo  = request.args.get("tipo",  "semanal")
    turno = request.args.get("turno", "todos")
    prestamos, fecha_desde, hoy = get_prestamos_informe(tipo, turno)
    return render_template("informe_resultado.html",
        prestamos=prestamos, tipo=tipo, turno=turno,
        fecha_desde=fecha_desde, hoy=hoy)

@app.route("/informes/excel")
def exportar_excel():
    if login_req(): return redirect(url_for("login"))
    try:
        import openpyxl
    except ImportError:
        return "Falta openpyxl. Ejecuta: pip install openpyxl", 500
    tipo  = request.args.get("tipo",  "semanal")
    turno = request.args.get("turno", "todos")
    rows, fecha_desde, hoy = get_prestamos_informe(tipo, turno)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prestamos"
    ws.append(["#","Herramienta","Estado","Usuario","Curso","Div",
                "Turno","Tipo","Sector","Destino","Prestamo","Devolucion","Obs"])
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"informe_{tipo}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── PLANTILLAS EXCEL ──────────────────────────────
@app.route("/plantillas/herramientas")
def plantilla_herramientas():
    if login_req(): return redirect(url_for("login"))
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Herramientas"
    ws.append(["nombre","descripcion","cantidad","estado"])
    ws.append(["Martillo","Martillo de carpintero",5,"Bueno"])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name="plantilla_herramientas.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/plantillas/usuarios")
def plantilla_usuarios():
    if login_req(): return redirect(url_for("login"))
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["nombre","curso","division","turno","tipo"])
    ws.append(["Juan Perez","3","A","Manana","Alumno"])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name="plantilla_usuarios.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    init_db()
    print("=" * 45)
    print("  Sistema Paniol CET 22 iniciado")
    print("  Abrir: http://127.0.0.1:5000/login")
    print("  Usuario: admin  |  Clave: cet22")
    print("=" * 45)
    app.run(debug=True, port=5000)